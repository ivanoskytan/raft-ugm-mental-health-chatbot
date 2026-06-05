from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions
from config.config import Settings
import io
import datetime
import logging
import os
import json

logger = logging.getLogger("FileService")
settings = Settings.load()
class FileService:

    CONTAINER_NAME = "assessments"

    @staticmethod
    def save_into_excel(assessment_map, chat_id):
        try: 
            wb = Workbook()

            default_sheet = wb.active
            wb.remove(default_sheet)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "..")
            list_mental_health_screening_path = os.path.join(BASE_DIR, "external_data", "original_mental_health_screening.json")
            
            scoring_lookup = {}
            if os.path.exists(list_mental_health_screening_path):
                with open(list_mental_health_screening_path, "r", encoding="utf-8") as f:
                    screening_data = json.load(f)
                    for group in screening_data:
                        scoring_lookup[group["section"]] = group.get("scoring_system", [])


            for section_data in assessment_map:
                section_name = section_data["section"]
                questions = section_data["questions"]

                sheet = wb.create_sheet(title=section_name)

                section_scoring = scoring_lookup.get(section_name, [])

                headers = ["No.", "Pertanyaan"] + [str(item["score"]) for item in section_scoring] 

                sheet.append(headers)

                for col in range(1, len(headers) + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align   
                    cell.border = thin_border   
                
                for idx, q in enumerate(questions, start=1):
                    row_data = [idx, q["original_question"]]
                    for item in section_scoring:
                        row_data.append(item["description"])
                    
                    sheet.append(row_data)

                    current_row = sheet.max_row
                    user_score = q.get("score")

                    for option_idx, item in enumerate(section_scoring, start=3):
                        cell = sheet.cell(row=current_row, column=option_idx)
                        if item["score"] == user_score:
                            cell.fill = yellow_fill

                sheet.column_dimensions["A"].width = 6
                sheet.column_dimensions["B"].width = 60

                for col_idx in range(3, len(headers) + 1):
                    col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(38 + col_idx)}"
                    sheet.column_dimensions[col_letter].width = 16

            file_name = f"Asesmen Kesehatan Mental_{chat_id}.xlsx"

            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            blob_service_client = BlobServiceClient.from_connection_string(settings.STORAGE_CONN_STR)
            container_client = blob_service_client.get_container_client(container=FileService.CONTAINER_NAME)
            if not container_client.exists():
                logger.info(f"Container '{FileService.CONTAINER_NAME}' does not exist. Creating container.")
                container_client.create_container()

            blob_client = blob_service_client.get_blob_client(container=FileService.CONTAINER_NAME, blob=file_name)
            logger.info("Uploading file to Azure Blob Storage at URL: " + blob_client.url)
            blob_client.upload_blob(excel_file, overwrite=True)

            sas_token = generate_blob_sas(
                account_name=blob_service_client.account_name,
                container_name=FileService.CONTAINER_NAME,
                blob_name=file_name,
                account_key=blob_service_client.credential.account_key,
                permission=BlobSasPermissions(read=True),
                expiry=datetime.datetime.utcnow() + datetime.timedelta(days=7)
            )

            sas_url = f"{blob_client.url}?{sas_token}"
            
            raw_bytes = excel_file.getvalue()

            return True, sas_url, raw_bytes, None
        
        except Exception as e:
            logger.error(f"Error saving assessment to Excel: {str(e)}")
            return False, None, f"[FileService]: Failed to create Excel {str(e)}"

